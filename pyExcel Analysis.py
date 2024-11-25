# Import necessary modules
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference


'''
Session: Lab 1X01
Group Members: Hans, Abizer & Yingwei
Due Date: November 24, 2023
Assignment 3: Data Science with Excel and Python
Summary: Create a Python program that takes an Excel spreadsheet of aircraft and animal collision data, and does the
following analysis, creating a new workbook, called aircraftWildlifeAnalysis.xlsx, with a sheet of information per
item: Determines with which animal the most collisions occurred, which year the most collisions occurred, which month
the most collisions occurred and which airline company had the most animal collisions.
Resources Used :
   Python Openpyxl, JavaTpoint, 2021. https://www.javatpoint.com/python-openpyxl
   https://www.softwaretestinghelp.com/python-openpyxl-tutorial/
   Python Excel, 2021. https://www.pythonexcel.com/openpyxl.php
'''
# Primary Contributor: Abizer
# Loads and Initializes the Workbooks
print("Loading workbook . . .")
wb = load_workbook("aircraftWildlifeStrikes.xlsx")
print("Workbook loaded. Here we go . . .")
wb1 = Workbook()
ws1 = wb1.active




# Primary Contributor: Hans
# Secondary Contributor: Yingwei
# Function for changing the animal names to their more common names
def animalNames(animal):
   animals = ["RAPTORS", "GULL", "DOVE", "SWALLOW", "OWL", "DEER", "SPARROW", "HAWK", "KESTREL", "LARK", "STARLING",
              "PIGEON", "GOOSE", "BLACKBIRD", "BAT", "PLOVER", "VULTURE", "DUCK", "SANDPIPER", "ROBIN",
              "MALLARD", "PERCHING BIRDS", "EGRET", "TERN", "SWIFT", "HERON", "THRUSH", "FALCON",
              "COYOTE", "CROW", "GRACKLE", "GEESE", "FINCH", "BUNTING", "OSPREY", "SKUNK", "RABBIT",
              "EAGLE", "COOT", "COWBIRD", "FLYCATCHER", "OPOSSUM", "WAXWING", "MARTIN", "VIREO",
              "WARBLER", "FOX", "SWAN", "PINTAIL", "WOODCHUCK", "JUNCO", "CORMORANT", "CRANE", "PIPIT",
              "CATBIRD", "MUNIA", "HARRIER", "MOCKINGBIRD", "RACCOON", "FLICKER", "MYNA",
              "SNIPE", "MERLIN", "WOODCOCK", "ELICAN", "YELLOWTHROAT", "PHEASANT", "KINGLET",
              "SHOVELER", "DUNLIN", "TURKEY", "GADWALL", "WIGEON"]  # A list of the most common animals in collisions
   for name in animals:  # Loops for every animal in the list
       if "," in animal:
           pass
       elif name in animal:  # If the animal is the same as the one in the list
           animal = name  # Changes the value of the animal to their more common name in the list
   return animal  # Returns the animal name




# Primary Contributor: Hans
# Secondary Contributor: Yingwei & Abizer
# Function for cleaning the column values and placing the proper values in a list
def cleanCol(col, animal):
   ws = wb.active
   for value in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=col, max_col=col, values_only=True):
       for cells in value:  # Loops for every cell in the values for the whole column read
           cell = str(cells)  # Makes the cell into a string for number errors
           if cell is None or cell == "None":  # Checks if the cell is empty
               pass  # Passes and does nothing
           elif "UNKNOWN" in cell:  # Checks if the cell is unknown
               pass
           else:  # If the value is not something unknown
               if animal == 1:  # Checks if the current loop is for animals, helps for run time
                   cell = animalNames(cell)
               dataList.append(cell)  # Adds the cell to the data list




# Primary Contributor: Hans
# Secondary Contributor: Yingwei
# Function for counting the amount of times the value happens in the list and places them into a dictionary
def counter():
   for item in dataList:  # Loops for every item in the list
       if item in dataDic:  # Checks if the item is in the dictionary
           dataDic[item] = dataDic[item] + 1  # Adds 1 to the dictionary value
       else:  # If the item is not in the dictionary
           dataDic[item] = 1  # Makes its value to 1
   return dataDic  # Returns the dictionary values




# Primary Contributor: Hans & Yingwei
# Secondary Contributor: Abizer
# Function for graphing the values into Excel
def graph(title, xtitle, ytitle):
   data = Reference(worksheet=ws1, min_col=1, min_row=2, max_row=len(orderedValues) + 2)  # Reads the data values
   val = Reference(worksheet=ws1, min_col=2, min_row=1, max_row=len(orderedValues) + 1)  # Reads the collision values
   chart = BarChart()  # Creates a chart
   chart.add_data(val, titles_from_data=True)  # Adds the collision values to graph
   chart.set_categories(data)  # Adds the data values to graph
   chart.height, chart.width = 10, 20 # Sets the size of the graph
   chart.title, chart.x_axis.title, chart.y_axis.title = title, xtitle, ytitle  # Labels the title and x and y-axis
   ws1.add_chart(chart, "E2")  # PLaces the chart in the location given
   wb1.save("aircraftWildlifeAnalysis.xlsx")  # Saves file in Excel




# Primary Contributor: Hans
# Loop for every category
for i in range(4):
   dataList = []  # Empty list for each loop
   dataDic = {}  # Empty dictionary for each loop
   if i == 0:  # Loop 1 for animals
       cleanCol(32, 1)  # Calls the cleaning function
       dataDic = counter()  # Calls the counting function and saves it to a variable
       if len(dataDic) > 15:  # Checks if there are more than 15 animals
           orderedValues = sorted(dataDic.items(), key=lambda x: x[1])[-15:]  # Sorts the dictionary according to
           # least to greatest number of collisions and limits the keys to 15 animals
           while orderedValues[0][1] < orderedValues[-1][1] * 0.1:  # Loops and checks the value of the animal
               orderedValues.pop(0)  # Gets rid of the data if the value is less than 10% of the highest value animal
       else:  # If there is less than 15 animals
           orderedValues = sorted(dataDic.items(), key=lambda x: x[1])  # Saves all values and orders them
       print("The animal(s) mostly involved in accidents are: ['" + str(orderedValues[-1][0]) + "'], with",
             orderedValues[-1][1], "incidents total")  # Prints message for the animal with the most collisions
       orderedValues = sorted(orderedValues)  # Reorders the values alphabetically
       ws1.append(("Animal", "Collisions"))  # Labels each column in the Excel Sheet
       ws1.title = "ChartForAnimals"  # Sets the title for the sheet
       for j in orderedValues:  # Loops for every value in the list
           ws1.append(j)  # PLaces values in the Excel Sheet
       graph("Aircraft and Animal Collisions by Animal Type", "Animal", "Quantity")
       # Graphs the values with the given titles and value location
   elif i == 1:  # Loop 2 for years
       cleanCol(2, 0)  # Similar concept as loop 1 with different values for some actions
       ws1 = wb1.active
       dataDic = counter()
       orderedValues = sorted(dataDic.items(), key=lambda x: x[1])  # No limiting value since all years are displayed
       print("The year(s) with most accidents are: [" + str(orderedValues[-1][0]) + "], with", orderedValues[-1][1],
             "incidents total")
       ws1 = wb1.create_sheet("ChartForYears")  # Creates a new sheet for the graph
       ws1.append(("Year", "Collisions"))
       orderedValues = sorted(orderedValues, key=lambda x: int(x[0]))  # Reorders the values numerically
       for j in orderedValues:
           ws1.append(j)
       graph("Aircraft and Animal Collisions by Animal Year", "Year", "Quantity")
   elif i == 2:  # Loop 3 for months
       cleanCol(3, 0)
       ws1 = wb1.active
       dataDic = counter()
       orderedValues = sorted(dataDic.items(), key=lambda x: x[1])  # No limiting value since all months are displayed
       print("The month(s) with most accidents are: [" + str(orderedValues[-1][0]) + "], with", orderedValues[-1][1],
             "incidents total")
       ws1 = wb1.create_sheet("ChartForMonths")
       ws1.append(("Month", "Collisions"))
       orderedValues = sorted(orderedValues, key=lambda x: int(x[0]))  # Reorders the values numerically
       for num in range(1, 13):  # Iterates numbers from 1-12
           if orderedValues[num - 1][0] != str(num):  # Checks if the value of the list is not the same as the range
               orderedValues.insert(num - 1, (str(num), 0))  # Inserts the empty valued month
       for j in orderedValues:
           ws1.append(j)
       graph("Aircraft and Animal Collisions by Animal Month", "Month", "Quantity")
   else:  # Loop 4 for airline companies
       cleanCol(6, 0)
       ws1 = wb1.active
       dataDic = counter()
       if len(dataDic) > 15:  # Checks if there are more than 15 Airline Companies
           orderedValues = sorted(dataDic.items(), key=lambda x: x[1])[-15:]  # Same concept as the animals
           while orderedValues[0][1] < orderedValues[-1][1] * 0.1:  # Limits the values to larger than 10% of max
               orderedValues.pop(0)
       else:
           orderedValues = sorted(dataDic.items(), key=lambda x: x[1])
       print("The airline(s) mostly involved in accidents are: ['" + str(orderedValues[-1][0]) + "'], with",
             orderedValues[-1][1], "incidents total")
       ws1 = wb1.create_sheet("ChartForAirlines")
       ws1.append(("Airline Company", "Collisions"))
       orderedValues = sorted(orderedValues)  # Reorders the values alphabetically
       for j in orderedValues:
           ws1.append(j)
       graph("Aircraft and Animal Collisions by Animal Airline Company", "Airline Company", "Quantity")